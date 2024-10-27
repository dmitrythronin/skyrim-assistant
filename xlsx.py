import openpyxl

REFLYEM_TABLE = 'reflyem.xlsx'
RFAD_TABLE = 'rfad.xlsx'


def count_words(text):
    """Подсчитывает количество слов в строке.

    Если входные данные не являются строкой или содержат 'коды крика', возвращает 0.
    """
    if not isinstance(text, str):  # Проверка, является ли текст строкой
        return 0
    if "Коды крика" in text:  # Проверка на наличие фразы 'коды крика'
        return 0
    return len(text.split())


def search_in_excel(file, search_string):
    # Загружаем книгу
    workbook = openpyxl.load_workbook(file)
    results = []

    # Проходим по всем листам в книге
    for sheet in workbook.worksheets:
        # Проходим по всем ячейкам в каждом листе
        for row in sheet.iter_rows():
            for cell in row:
                # Проверяем, содержит ли ячейка строку
                if isinstance(cell.value, str) and search_string.lower() in cell.value.lower().replace('"', ''):
                    comment_text = cell.comment.text if cell.comment else None

                    # Проверяем на пустоту комментарий и ищем альтернативные значения
                    if not comment_text or count_words(comment_text) < 5:
                        found = False  # Флаг, указывающий на нахождение подходящей ячейки

                        # Проверяем ячейки ниже на пустоту и количество слов
                        for i in [1, 2, 3, 5, 6, 7]:  # проверяем на 1, 2, 3 и 7 ячеек вниз
                            next_cell = sheet.cell(row=cell.row + i, column=cell.column)
                            if next_cell.value:
                                # Проверяем количество слов в следующей ячейке
                                if count_words(next_cell.value) >= 5:
                                    comment_text = next_cell.value
                                    found = True
                                    break  # Выходим из цикла, если нашли подходящее значение

                        # Если совпадений всё ещё нет, проверяем ячейку вправо
                        if not found:
                            right_cell = sheet.cell(row=cell.row, column=cell.column + 1)
                            if right_cell.value and count_words(right_cell.value) >= 5:
                                comment_text = right_cell.value

                    results.append((sheet.title, cell.coordinate, cell.value, comment_text))

    return results


def search_table(table, search_string):
    found_data = search_in_excel(table, search_string)

    # Вывод результатов
    if found_data:
        for sheet_name, cell_coord, cell_value, comment in found_data:
            if comment:
                return f'{cell_value}\n{comment}'
    return "Совпадений не найдено."
